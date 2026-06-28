"""
NSE Stockbee Scanner — Report Generator
=========================================
Generates:
  1. Daily Excel Report (multi-sheet workbook)
       Sheet 1 — Today's Signals (MB + EP, sorted by score)
       Sheet 2 — Anticipation Watchlist
       Sheet 3 — Market Monitor Summary
       Sheet 4 — Backtest / Performance Summary
  2. Telegram daily signal cards (concise, mobile-friendly)
  3. Weekly breadth report (text summary for study)
"""

from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter

from config import REPORTS_DIR, MB_SCORE_THRESHOLDS, EP_SCORE_THRESHOLDS
from logger_utils import get_logger
from market_monitor import MarketMonitorSnapshot
from stockbee_scanner import StockbeeSignal

log = get_logger("scanner")

# ─── Colour palette ───────────────────────────────────────────────────────────
CLR = {
    "header_bg":   "1F4E79",  # dark navy
    "header_fg":   "FFFFFF",
    "elite":       "00B050",  # dark green
    "strong":      "92D050",  # light green
    "watch":       "FFEB9C",  # amber
    "weak":        "FFC7CE",  # light red
    "alt_row":     "EBF3FB",  # light blue
    "white":       "FFFFFF",
    "border":      "BDD7EE",
    "title_bg":    "2E75B6",  # mid blue
    "title_fg":    "FFFFFF",
    "bull_bg":     "E2EFDA",  # light green bg
    "bear_bg":     "FCE4D6",  # light red bg
    "caution_bg":  "FFF2CC",  # amber bg
    "neutral_bg":  "EDEDED",  # grey bg
}

THIN   = Side(style="thin",   color=CLR["border"])
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ─── Main entry point ─────────────────────────────────────────────────────────

def generate_report(
    signals:         List[StockbeeSignal],
    market_snapshot: MarketMonitorSnapshot,
    today:           Optional[date] = None,
) -> Path:
    """
    Build the Excel workbook and return its path.
    signals = full list of StockbeeSignal objects from today's scan.
    """
    today = today or date.today()
    out_path = REPORTS_DIR / f"stockbee_report_{today.isoformat()}.xlsx"

    wb = Workbook()
    wb.remove(wb.active)

    # Separate MB and EP signals
    mb_signals  = [s for s in signals if "MB_BREAKOUT" == s.signal_type and s.tradeable]
    ant_signals = [s for s in signals if "MB_ANTICIPATION" == s.signal_type]
    ep_signals  = [s for s in signals if s.signal_type.startswith("EP")]

    _sheet_signals(wb, mb_signals, ep_signals, today)
    _sheet_anticipation(wb, ant_signals, today)
    _sheet_market_monitor(wb, market_snapshot, today)
    _sheet_avoid(wb, signals, today)

    wb.save(out_path)
    log.info("Report saved: %s", out_path)
    return out_path


# ─── Sheet 1: Today's Signals ─────────────────────────────────────────────────

def _sheet_signals(
    wb:          Workbook,
    mb_signals:  List[StockbeeSignal],
    ep_signals:  List[StockbeeSignal],
    today:       date,
) -> None:
    ws = wb.create_sheet("📈 Signals Today")
    ws.sheet_view.showGridLines = False

    _write_title(ws, f"NSE Stockbee Scanner — {today.strftime('%d %b %Y')}  |  SIGNALS TODAY", row=1, colspan=16)

    # ── Section A: MB Breakouts ───────────────────────────────────────────────
    _write_section_header(ws, "SECTION A — Momentum Burst Breakouts (Exit Day 3 partial / Day 5 full)", row=3, colspan=16)

    MB_HEADERS = [
        "Symbol", "Score", "Class", "RS Rank", "TI65",
        "2LYNCH", "Y?", "Entry ₹", "Stop ₹", "Stop%",
        "Day3 T ₹", "Day5 T ₹", "Brkout%", "Vol Ratio",
        "Consol Bars", "Consol Width%"
    ]
    _write_header_row(ws, MB_HEADERS, row=4)

    row = 5
    for sig in mb_signals:
        s = sig.setup
        stop_pct = round((sig.entry_price - sig.stop_loss) / sig.entry_price * 100, 1)
        lynch_str = _lynch_flag_str(s.twolynch_flags)
        row_data = [
            sig.symbol.replace(".NS", ""),
            sig.composite_score,
            sig.classification,
            s.rs_rank,
            s.ti65,
            f"{s.twolynch_score}/5  {lynch_str}",
            "✅" if s.is_young_trend else "❌",
            sig.entry_price,
            sig.stop_loss,
            f"{stop_pct}%",
            sig.target_1,
            sig.target_2,
            f"{s.breakout_pct}%",
            s.volume_ratio,
            s.consolidation_bars,
            f"{s.consolidation_width_pct}%",
        ]
        _write_data_row(ws, row_data, row, alt=(row % 2 == 0),
                        class_col=3, classification=sig.classification)
        row += 1

    if not mb_signals:
        ws.cell(row=row, column=1).value = "No MB Breakout signals today"
        row += 1

    row += 1

    # ── Section B: EP Signals ─────────────────────────────────────────────────
    _write_section_header(ws, "SECTION B — Episodic Pivot Signals (Hold up to 30 days)", row=row, colspan=16)
    row += 1

    EP_HEADERS = [
        "Symbol", "EP Type", "Score", "Class", "RS Rank",
        "Gap%", "Day Chg%", "Vol Spike", "Quiet Days",
        "Entry ₹", "Stop ₹", "Stop%", "+20% T", "+40% T", "Neglect", "Conv"
    ]
    _write_header_row(ws, EP_HEADERS, row=row)
    row += 1

    for sig in ep_signals:
        ep = sig.setup
        stop_pct = round((sig.entry_price - sig.stop_loss) / sig.entry_price * 100, 1)
        row_data = [
            sig.symbol.replace(".NS", ""),
            ep.ep_type.replace("EP_", ""),
            sig.composite_score,
            sig.classification,
            ep.rs_rank,
            f"{ep.gap_pct}%",
            f"{ep.day_change_pct}%",
            f"{ep.volume_spike_ratio}x",
            ep.prior_quiet_days,
            sig.entry_price,
            sig.stop_loss,
            f"{stop_pct}%",
            sig.target_1,
            sig.target_2,
            round(ep.neglect_score, 2),
            "🔥 HIGH" if ep.is_high_conviction else "normal",
        ]
        _write_data_row(ws, row_data, row, alt=(row % 2 == 0),
                        class_col=4, classification=sig.classification)
        row += 1

    if not ep_signals:
        ws.cell(row=row, column=1).value = "No EP signals today"
        row += 1

    _auto_col_width(ws)


# ─── Sheet 2: Anticipation Watchlist ─────────────────────────────────────────

def _sheet_anticipation(
    wb:          Workbook,
    ant_signals: List[StockbeeSignal],
    today:       date,
) -> None:
    ws = wb.create_sheet("🔭 Anticipation")
    ws.sheet_view.showGridLines = False

    _write_title(ws, "ANTICIPATION WATCHLIST — Pre-Breakout Coiling Setups", row=1, colspan=14)
    ws.cell(row=2, column=1).value = (
        "These stocks are in consolidation and may break out within 1-5 days. "
        "Monitor daily. Enter only when 4%+ breakout occurs with volume."
    )
    ws.cell(row=2, column=1).font = Font(italic=True, size=9)

    HEADERS = [
        "Symbol", "Score", "Class", "RS Rank", "TI65",
        "2LYNCH", "Consol Bars", "Width%", "Vol Dry%",
        "Prior Move%", "Watch Entry ₹", "Stop ₹", "Watch Zone Low", "Watch Zone High"
    ]
    _write_header_row(ws, HEADERS, row=4)

    row = 5
    for sig in ant_signals:
        s = sig.setup
        vol_dry = round((1 - s.volume_ratio) * 100, 1)   # approximate
        row_data = [
            sig.symbol.replace(".NS", ""),
            sig.composite_score,
            sig.classification,
            s.rs_rank,
            s.ti65,
            f"{s.twolynch_score}/5",
            s.consolidation_bars,
            f"{s.consolidation_width_pct}%",
            f"{vol_dry}%",
            f"{s.prior_move_pct}%",
            sig.entry_price,
            sig.stop_loss,
            round(sig.entry_price * 0.995, 2),
            round(sig.entry_price * 1.01, 2),
        ]
        _write_data_row(ws, row_data, row, alt=(row % 2 == 0),
                        class_col=3, classification=sig.classification)
        row += 1

    if not ant_signals:
        ws.cell(row=row, column=1).value = "No anticipation setups today"

    _auto_col_width(ws)


# ─── Sheet 3: Market Monitor ──────────────────────────────────────────────────

def _sheet_market_monitor(
    wb:       Workbook,
    mm:       MarketMonitorSnapshot,
    today:    date,
) -> None:
    ws = wb.create_sheet("📊 Market Monitor")
    ws.sheet_view.showGridLines = False

    regime_bg = {
        "BULL": CLR["bull_bg"], "NEUTRAL": CLR["neutral_bg"],
        "CAUTION": CLR["caution_bg"], "BEAR": CLR["bear_bg"],
    }.get(mm.market_regime, CLR["neutral_bg"])

    _write_title(ws, f"MARKET MONITOR — {today}  |  Regime: {mm.market_regime}  (Score: {mm.regime_score}/100)",
                 row=1, colspan=4, bg=regime_bg, fg="000000")

    metrics = [
        ("EMA BREADTH", ""),
        ("% Above 200 EMA", f"{mm.pct_above_200ema:.1f}%",
         "✅" if mm.pct_above_200ema >= 60 else "🔴" if mm.pct_above_200ema < 40 else "🟡"),
        ("% Above 50 EMA", f"{mm.pct_above_50ema:.1f}%", ""),
        ("% Above 20 EMA", f"{mm.pct_above_20ema:.1f}%", ""),
        ("", ""),
        ("DAILY MOMENTUM", ""),
        ("Stocks Up 4%+", str(mm.up_4pct_count),
         "✅" if mm.up_4pct_count >= 15 else "⚠️" if mm.up_4pct_count >= 5 else "🔴"),
        ("Stocks Down 4%+", str(mm.down_4pct_count),
         "🔴" if mm.down_4pct_count >= 30 else "⚠️" if mm.down_4pct_count >= 15 else "✅"),
        ("Advance / Decline", f"{mm.advance_count} / {mm.decline_count}  ({mm.advance_decline_ratio:.2f})",
         "✅" if mm.advance_decline_ratio >= 1.5 else "🔴" if mm.advance_decline_ratio < 0.5 else "🟡"),
        ("", ""),
        ("52-WEEK BREADTH", ""),
        ("% at 52W Highs", f"{mm.pct_52w_highs:.1f}%",
         "✅" if mm.pct_52w_highs >= 3 else "⚠️"),
        ("% at 52W Lows", f"{mm.pct_52w_lows:.1f}%",
         "🔴" if mm.pct_52w_lows >= 2 else "✅"),
        ("", ""),
        ("TI65 BREADTH", ""),
        ("TI65 Green Count", f"{mm.ti65_green_count}  ({mm.ti65_green_pct:.1f}%)",
         "✅" if mm.ti65_green_pct >= 40 else "🟡"),
        ("", ""),
        ("WEEKLY MOMENTUM", ""),
        ("Stocks Up 20%+ (5d)", str(mm.weekly_up20_count), ""),
        ("Stocks Down 20%+ (5d)", str(mm.weekly_down20_count), ""),
        ("Universe Size", str(mm.universe_size), ""),
        ("", ""),
        ("TRADING GUIDANCE", ""),
        ("Regime Score", f"{mm.regime_score:.0f} / 100", ""),
        ("Buy Breakouts Aggressively?",
         "YES — FFM is ON 🟢" if mm.buy_breakouts_aggressively else "NO",
         ""),
        ("Long Setups Allowed?",
         "YES" if mm.trading_allowed else "NO — study only 🚫", ""),
        ("EP Setups Allowed?",
         "YES (catalyst overrides)" if mm.market_regime in ("BULL", "NEUTRAL", "CAUTION") else "NO", ""),
    ]

    row = 3
    for item in metrics:
        if item[0] == "":
            row += 1
            continue
        if item[1] == "":  # section header
            ws.cell(row=row, column=1).value = item[0]
            ws.cell(row=row, column=1).font = Font(bold=True, color="2E75B6")
            row += 1
            continue

        ws.cell(row=row, column=1).value = item[0]
        ws.cell(row=row, column=2).value = item[1]
        if len(item) > 2:
            ws.cell(row=row, column=3).value = item[2]
        ws.cell(row=row, column=1).font = Font(size=10)
        ws.cell(row=row, column=2).font = Font(size=10, bold=True)
        if row % 2 == 0:
            for col in range(1, 4):
                ws.cell(row=row, column=col).fill = PatternFill(
                    fill_type="solid", fgColor=CLR["alt_row"])
        row += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 10


# ─── Sheet 4: Stocks to Avoid ─────────────────────────────────────────────────

def _sheet_avoid(
    wb:      Workbook,
    signals: List[StockbeeSignal],
    today:   date,
) -> None:
    ws = wb.create_sheet("⚠️ Low Quality")
    ws.sheet_view.showGridLines = False

    _write_title(ws, "LOW QUALITY / AVOID — Signals That Failed Quality Filters", row=1, colspan=8)
    ws.cell(row=2, column=1).value = (
        "These signals fired the primary scan but failed 2LYNCH, regime, or quality filters. "
        "Study why — this is how you improve pattern recognition."
    )
    ws.cell(row=2, column=1).font = Font(italic=True, size=9)

    weak = [s for s in signals if s.classification == "Weak" or not s.tradeable]
    HEADERS = ["Symbol", "Signal Type", "Score", "Regime", "RS Rank", "TI65", "2LYNCH", "Reason"]
    _write_header_row(ws, HEADERS, row=4)

    row = 5
    for sig in weak[:50]:  # cap at 50
        s  = sig.setup
        rs = getattr(s, "rs_rank", "-")
        ti = getattr(s, "ti65", "-")
        lynch = getattr(s, "twolynch_score", "-")
        reason = "Not tradeable" if not sig.tradeable else "Low score"
        _write_data_row(ws, [
            sig.symbol.replace(".NS", ""),
            sig.signal_type,
            sig.composite_score,
            sig.market_snapshot.market_regime,
            rs, ti, lynch, reason,
        ], row, alt=(row % 2 == 0))
        row += 1

    if not weak:
        ws.cell(row=row, column=1).value = "All signals passed quality filters today ✅"

    _auto_col_width(ws)


# ─── Telegram signal cards ─────────────────────────────────────────────────────

def format_telegram_signals(
    signals:  List[StockbeeSignal],
    snapshot: MarketMonitorSnapshot,
    max_cards: int = 10,
) -> List[str]:
    """
    Returns list of Telegram-ready message strings.
    One string per signal card, plus one Market Monitor header card.
    """
    messages: List[str] = []

    # Card 0: Market Monitor header
    regime_emoji = {
        "BULL": "🟢", "NEUTRAL": "🟡",
        "CAUTION": "🟠", "BEAR": "🔴"
    }.get(snapshot.market_regime, "⚪")

    mm_card = (
        f"📊 *MARKET MONITOR — {snapshot.snapshot_date}*\n"
        f"Regime: {regime_emoji} *{snapshot.market_regime}* (Score: {snapshot.regime_score:.0f}/100)\n"
        f"Above 200 EMA: {snapshot.pct_above_200ema:.1f}%\n"
        f"Up 4%+: {snapshot.up_4pct_count} | Down 4%+: {snapshot.down_4pct_count}\n"
        f"A/D Ratio: {snapshot.advance_decline_ratio:.2f}\n"
        f"{'✅ BUY AGGRESSIVELY' if snapshot.buy_breakouts_aggressively else '⚠️ BE SELECTIVE' if snapshot.trading_allowed else '🚫 AVOID LONGS'}"
    )
    messages.append(mm_card)

    # Signal cards (top N tradeable, sorted by score)
    tradeable = [s for s in signals if s.tradeable][:max_cards]

    for sig in tradeable:
        s  = sig.setup
        rs = getattr(s, "rs_rank", "?")
        ti = getattr(s, "ti65", "?")

        if sig.signal_type == "MB_BREAKOUT":
            lynch_score = getattr(s, "twolynch_score", "?")
            lynch_flags = getattr(s, "twolynch_flags", {})
            stop_pct    = round((sig.entry_price - sig.stop_loss) / sig.entry_price * 100, 1)
            card = (
                f"📈 *MB BREAKOUT — {sig.symbol.replace('.NS', '')}*\n"
                f"Score: {sig.composite_score:.0f}/100 | Class: *{sig.classification}*\n"
                f"─────────────────\n"
                f"Entry: ₹{sig.entry_price} | Stop: ₹{sig.stop_loss} ({stop_pct}%)\n"
                f"Day 3 T: ₹{sig.target_1} | Day 5 T: ₹{sig.target_2}\n"
                f"─────────────────\n"
                f"RS: {rs} | TI65: {ti} | 2LYNCH: {lynch_score}/5\n"
                f"{_lynch_flag_str(lynch_flags)}\n"
                f"Breakout: {getattr(s, 'breakout_pct', '?')}% | Vol Ratio: {getattr(s, 'volume_ratio', '?')}x\n"
                f"Consol: {getattr(s, 'consolidation_bars', '?')} bars @ {getattr(s, 'consolidation_width_pct', '?')}% wide\n"
                f"Prior Move: {getattr(s, 'prior_move_pct', '?')}% | Young: {'✅' if getattr(s, 'is_young_trend', False) else '❌'}\n"
                f"Regime: {sig.market_snapshot.market_regime}\n"
                f"⏰ Hold max {sig.max_hold_days} days"
            )

        elif sig.signal_type == "MB_ANTICIPATION":
            stop_pct = round((sig.entry_price - sig.stop_loss) / sig.entry_price * 100, 1)
            card = (
                f"🔭 *ANTICIPATION — {sig.symbol.replace('.NS', '')}*\n"
                f"Score: {sig.composite_score:.0f}/100 | Class: *{sig.classification}*\n"
                f"─────────────────\n"
                f"Watch Entry: ₹{sig.entry_price} | Stop: ₹{sig.stop_loss} ({stop_pct}%)\n"
                f"⚠️ Enter ONLY on 4%+ breakout with volume\n"
                f"─────────────────\n"
                f"RS: {rs} | TI65: {ti} | 2LYNCH: {getattr(s, 'twolynch_score', '?')}/5\n"
                f"Consol: {getattr(s, 'consolidation_bars', '?')} bars | "
                f"Width: {getattr(s, 'consolidation_width_pct', '?')}%\n"
                f"Prior Move: {getattr(s, 'prior_move_pct', '?')}%"
            )

        else:
            # EP signal
            ep = s
            stop_pct = round((sig.entry_price - sig.stop_loss) / sig.entry_price * 100, 1)
            ep_label = sig.signal_type.replace("EP_", "")
            card = (
                f"🚀 *{ep_label} EP — {sig.symbol.replace('.NS', '')}*\n"
                f"Score: {sig.composite_score:.0f}/100 | Class: *{sig.classification}*\n"
                f"{'🔥 HIGH CONVICTION' if getattr(ep, 'is_high_conviction', False) else ''}\n"
                f"─────────────────\n"
                f"Entry: ₹{sig.entry_price} | Stop: ₹{sig.stop_loss} ({stop_pct}%)\n"
                f"+20%: ₹{sig.target_1} | +40%: ₹{sig.target_2}\n"
                f"─────────────────\n"
                f"Gap: {getattr(ep, 'gap_pct', '?')}% | Day Chg: {getattr(ep, 'day_change_pct', '?')}%\n"
                f"Vol Spike: {getattr(ep, 'volume_spike_ratio', '?')}x | "
                f"Quiet Days: {getattr(ep, 'prior_quiet_days', '?')}\n"
                f"Neglect: {getattr(ep, 'neglect_score', '?')} | RS: {rs}\n"
                f"Regime: {sig.market_snapshot.market_regime}\n"
                f"⏰ Hold up to {sig.max_hold_days} days, trail after +20%"
            )

        messages.append(card)

    return messages


def format_weekly_report(weekly_data: dict) -> str:
    """Format the weekly breadth scan for Telegram / text report."""
    lines = [
        "📅 *WEEKLY STUDY REPORT*",
        f"Generated: {date.today()}",
        "",
        "📈 *Top 10 Weekly Winners (Up 20%+ in 5 days)*",
    ]
    for r in weekly_data.get("weekly_winners", [])[:10]:
        lines.append(f"  {r['symbol'].replace('.NS', ''):<20} +{r['wk_chg']:.1f}%  ₹{r['close']}")

    lines += ["", "📉 *Top 10 Weekly Losers (Down 20%+ in 5 days)*"]
    for r in weekly_data.get("weekly_losers", [])[:10]:
        lines.append(f"  {r['symbol'].replace('.NS', ''):<20} {r['wk_chg']:.1f}%  ₹{r['close']}")

    lines += ["", "🏆 *Top Monthly Leaders (Up 50%+ in 40 days)*"]
    for r in weekly_data.get("monthly_leaders", [])[:15]:
        lines.append(f"  {r['symbol'].replace('.NS', ''):<20} +{r['mo_chg']:.1f}%  ₹{r['close']}")

    lines += ["", "─── Study each of these ───"]
    lines.append("Ask: EP day visible? Clean consolidation? MB setups after EP?")

    return "\n".join(lines)


# ─── Shared worksheet helpers ─────────────────────────────────────────────────

def _write_title(ws, text: str, row: int, colspan: int, bg: str = CLR["title_bg"], fg: str = CLR["title_fg"]):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=colspan)
    cell = ws.cell(row=row, column=1)
    cell.value = text
    cell.font  = Font(bold=True, color=fg, size=12)
    cell.fill  = PatternFill(fill_type="solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 22


def _write_section_header(ws, text: str, row: int, colspan: int):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=colspan)
    cell = ws.cell(row=row, column=1)
    cell.value = text
    cell.font  = Font(bold=True, color=CLR["header_fg"], size=10)
    cell.fill  = PatternFill(fill_type="solid", fgColor="2E75B6")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 18


def _write_header_row(ws, headers: list, row: int):
    for col, hdr in enumerate(headers, 1):
        c = ws.cell(row=row, column=col)
        c.value = hdr
        c.font  = Font(bold=True, color=CLR["header_fg"], size=9)
        c.fill  = PatternFill(fill_type="solid", fgColor=CLR["header_bg"])
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[row].height = 18


def _write_data_row(
    ws, row_data: list, row: int, alt: bool = False,
    class_col: Optional[int] = None, classification: Optional[str] = None
):
    bg = CLR["alt_row"] if alt else CLR["white"]
    fill = PatternFill(fill_type="solid", fgColor=bg)

    for col, val in enumerate(row_data, 1):
        c = ws.cell(row=row, column=col)
        c.value  = val
        c.font   = Font(size=9)
        c.border = BORDER
        c.fill   = fill
        c.alignment = Alignment(horizontal="center")

    # Colour classification cell
    if class_col and classification:
        class_fill = {
            "Elite":  CLR["elite"],
            "Strong": CLR["strong"],
            "Watch":  CLR["watch"],
            "Weak":   CLR["weak"],
        }.get(classification, CLR["white"])
        cc = ws.cell(row=row, column=class_col)
        cc.fill = PatternFill(fill_type="solid", fgColor=class_fill)
        cc.font = Font(bold=True, size=9)


def _auto_col_width(ws, min_w: int = 8, max_w: int = 30):
    for col_cells in ws.columns:
        length = max(
            (len(str(cell.value)) for cell in col_cells if cell.value), default=min_w
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(
            max(length + 2, min_w), max_w
        )


def _lynch_flag_str(flags: dict) -> str:
    if not flags:
        return ""
    mapping = {
        "2_not_up_two_days":    "2",
        "L_linear_prior_move":  "L",
        "Y_young_trend":        "Y",
        "N_narrow_or_negative": "N",
        "C_clean_consolidation":"C",
        "H_close_near_high":    "H",
    }
    passed = [short for key, short in mapping.items() if flags.get(key) is True]
    failed = [short for key, short in mapping.items() if flags.get(key) is False]
    return f"✅{''.join(passed)} ❌{''.join(failed)}" if failed else f"✅{''.join(passed)}"
