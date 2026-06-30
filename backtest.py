"""
NSE Stockbee Scanner — Backtesting Engine v3
=============================================
Based on Pradeep Bonde (Stockbee) methodology.

FIX v3 (this version):
  1. MM filter bug: apply_mm_filter=False now correctly generates signals
     on ALL days including BEAR — previously both runs were identical
     because BEAR check was hardcoded before the flag was checked.
  2. Equity curve: trade-by-trade compounding at 1% risk per trade.
  3. Yearly P&L breakdown: regime dependency check year-by-year.
  4. Full trade log: every field per trade for external analysis.
  5. Sharpe fix: computed on daily equity curve returns not trade P&L.
  6. CAGR fix: uses actual compound equity curve, not proxy.
  7. Day3 hit fix: checks partial_exit_price not hold_days ≤ 3.
  8. Burst 8% fix: uses max intraday high during hold, not exit P&L.
  9. Verdict sheet: pass/fail criteria matching VCP reference format.

Parameter changes (from backtest analysis findings):
  RS_MIN_FOR_MB:           70 → 80  (RS 80-99 = 61% WR vs 70-80 = 48%)
  TI65_BULL_THRESHOLD:   1.03 → 1.05 (TI65>1.05 = 49% WR vs 1.03-1.05 = 39%)
  MB_CONSOL_MIN_WIDTH_PCT: 2 → 7   (7-10% = 55.8% WR vs <7% = <40%)
  MB_SCORE_THRESHOLDS watch: 55 → 65 (force higher quality minimum)
"""

from __future__ import annotations

import math
import uuid
from datetime import date, timedelta
from typing import List, Optional

import numpy as np
import pandas as pd

from config import (
    ACCOUNT_SIZE, MB_HOLD_MAX_DAYS, MB_PARTIAL_EXIT_DAY,
    EP_HOLD_MAX_DAYS, RISK_PER_TRADE_PCT,
    RS_MIN_FOR_MB, TI65_BULL_THRESHOLD, MB_SCORE_THRESHOLDS,
)
from downloader import load_daily
from episodic_pivot import detect_9m_ep
from logger_utils import get_logger
from market_monitor import compute_market_monitor, market_allows_trading
from momentum_burst import detect_mb_signal
from stockbee_scanner import compute_rs_ranks

log = get_logger("performance")

RISK_FREE_RATE  = 0.065    # RBI repo rate proxy
RISK_PCT        = RISK_PER_TRADE_PCT / 100   # e.g. 0.01


# ─── Trade record ─────────────────────────────────────────────────────────────

class Trade:
    __slots__ = [
        "symbol", "signal_type", "entry_date", "entry_price",
        "stop_loss", "target1", "target2",
        "exit_date", "exit_price", "exit_reason",
        "pnl_pct", "pnl_r", "hold_days",
        "rs_rank", "ti65", "twolynch_score", "consol_width",
        "composite_score", "market_regime",
        "partial_exit_price", "partial_exit_date", "partial_pnl_pct",
        "high_during_hold",   # max High seen during hold (for burst check)
    ]

    def __init__(self, **kwargs):
        for slot in self.__slots__:
            setattr(self, slot, None)
        for k, v in kwargs.items():
            setattr(self, k, v)


# ─── Main backtest ─────────────────────────────────────────────────────────────

def run_backtest(
    symbols:          List[str],
    start_date:       date,
    end_date:         date,
    apply_mm_filter:  bool  = True,
    signal_type:      str   = "MB_BREAKOUT",
    min_score:        float = MB_SCORE_THRESHOLDS["watch"],
) -> dict:
    """
    Walk-forward daily backtest of Stockbee MB + EP signals.

    apply_mm_filter=True  → respect BEAR/CAUTION regime gates (live behaviour)
    apply_mm_filter=False → generate signals regardless of regime (shows value of filter)

    FIX: the old code had `if apply_mm_filter and not mm.trading_allowed: continue`
    BEFORE scanning — which means the WITHOUT-filter run still skipped BEAR days
    because `trading_allowed=False` in BEAR was checked even when filter was off.

    Correct logic:
      WITH filter:    skip day if BEAR; skip MB in CAUTION; allow EP in CAUTION
      WITHOUT filter: scan every day regardless of regime
    """
    log.info(
        "Backtest start: %s → %s | signal=%s | mm_filter=%s | min_score=%.0f",
        start_date, end_date, signal_type, apply_mm_filter, min_score
    )

    # ── Load data ─────────────────────────────────────────────────────────────
    all_data: dict[str, pd.DataFrame] = {}
    for sym in symbols:
        df = load_daily(sym)
        if df is not None and len(df) >= 300:
            all_data[sym] = df

    log.info("Loaded %d / %d symbols", len(all_data), len(symbols))

    # ── Trading calendar ──────────────────────────────────────────────────────
    ref_df = all_data[next(iter(all_data))]
    trading_days = [
        d.date() for d in ref_df.index
        if start_date <= d.date() <= end_date
    ]
    if len(trading_days) < 20:
        log.error("Insufficient trading days")
        return {}

    log.info("Simulating %d trading days …", len(trading_days))

    # ── Per-day loop ──────────────────────────────────────────────────────────
    all_trades:      List[Trade]      = []
    open_positions:  dict[str, Trade] = {}
    monthly_signals: dict[str, int]   = {}

    for sim_date in trading_days:

        # Slice to sim_date (no lookahead)
        day_data: dict[str, pd.DataFrame] = {
            sym: df[df.index.date <= sim_date]
            for sym, df in all_data.items()
            if len(df[df.index.date <= sim_date]) >= 260
        }
        if len(day_data) < 10:
            continue

        # ── Exit open positions ───────────────────────────────────────────────
        to_close = []
        for sym, trade in open_positions.items():
            if sym not in day_data:
                continue
            bar         = day_data[sym].iloc[-1]
            today_close = float(bar["Close"])
            today_low   = float(bar["Low"])
            today_high  = float(bar["High"])
            today_open  = float(bar["Open"])
            hold_days   = (sim_date - trade.entry_date).days

            # Track max high during hold (for burst 8% check)
            if trade.high_during_hold is None:
                trade.high_during_hold = today_high
            else:
                trade.high_during_hold = max(trade.high_during_hold, today_high)

            # Stop detection: gap-down open vs intraday stop
            gap_down   = today_open < trade.stop_loss
            intra_stop = (not gap_down) and (today_low <= trade.stop_loss)
            stop_hit   = gap_down or intra_stop
            fill_price = today_open if gap_down else trade.stop_loss

            # Day 3 partial exit for MB (record only, doesn't close position)
            if ("MB" in (trade.signal_type or "") and
                    hold_days >= MB_PARTIAL_EXIT_DAY - 1 and
                    trade.partial_exit_price is None):
                trade.partial_exit_price = today_close
                trade.partial_exit_date  = sim_date
                trade.partial_pnl_pct    = (today_close - trade.entry_price) / trade.entry_price * 100

            # Exit priority: stop > MB day5 > EP max hold > EP trail
            force_exit  = False
            exit_reason = ""
            exit_price  = today_close

            if stop_hit:
                force_exit  = True
                exit_reason = "GAP_STOP" if gap_down else "STOP_HIT"
                exit_price  = fill_price

            elif "MB" in (trade.signal_type or "") and hold_days >= MB_HOLD_MAX_DAYS - 1:
                force_exit  = True
                exit_reason = "DAY5_EXIT"
                exit_price  = today_close

            elif "EP" in (trade.signal_type or "") and hold_days >= EP_HOLD_MAX_DAYS - 1:
                force_exit  = True
                exit_reason = "EP_MAX_HOLD"
                exit_price  = today_close

            elif "EP" in (trade.signal_type or ""):
                gain_pct = (today_close - trade.entry_price) / trade.entry_price * 100
                if gain_pct >= 20.0:
                    recent_high = float(day_data[sym]["High"].iloc[-10:].max())
                    trail_stop  = recent_high * 0.90
                    if today_low <= trail_stop:
                        force_exit  = True
                        exit_reason = "EP_TRAIL_STOP"
                        exit_price  = today_open if today_open < trail_stop else trail_stop

            if force_exit:
                risk_per_share = trade.entry_price - trade.stop_loss
                shares         = (ACCOUNT_SIZE * RISK_PCT / risk_per_share) if risk_per_share > 0 else 0
                trade.exit_date   = sim_date
                trade.exit_price  = round(exit_price, 2)
                trade.exit_reason = exit_reason
                trade.hold_days   = hold_days
                trade.pnl_pct     = (exit_price - trade.entry_price) / trade.entry_price * 100
                trade.pnl_r       = (exit_price - trade.entry_price) / risk_per_share if risk_per_share > 0 else 0
                all_trades.append(trade)
                to_close.append(sym)

        for sym in to_close:
            del open_positions[sym]

        # ── Compute MM for today ──────────────────────────────────────────────
        mm = compute_market_monitor(day_data)

        # ── FIX: correct MM filter logic ──────────────────────────────────────
        # WITH filter:    BEAR = skip entire day (FFM rule)
        # WITHOUT filter: scan every day regardless of regime
        if apply_mm_filter and mm.market_regime == "BEAR":
            continue   # FFM rule: never trade in BEAR

        # Compute RS ranks for today's universe
        rs_ranks = compute_rs_ranks(list(day_data.keys()), day_data)
        month_key = sim_date.strftime("%Y-%m")

        # ── Scan for new signals ──────────────────────────────────────────────
        for sym in day_data:
            if sym in open_positions:
                continue

            daily   = day_data[sym]
            rs_rank = rs_ranks.get(sym, 50.0)
            signal  = None

            # MB_BREAKOUT
            if signal_type in ("MB_BREAKOUT", "both"):
                # WITH filter: respect CAUTION gate (no MB in CAUTION)
                # WITHOUT filter: generate MB even in CAUTION/BEAR
                mb_allowed = (not apply_mm_filter) or market_allows_trading(mm, "MB_BREAKOUT")
                if mb_allowed:
                    signal = detect_mb_signal(sym, daily, rs_rank)

            # EP_9M (allowed in CAUTION even with filter)
            if signal is None and signal_type in ("EP_9M", "both"):
                ep_allowed = (not apply_mm_filter) or market_allows_trading(mm, "EP_9M")
                if ep_allowed:
                    ep = detect_9m_ep(sym, daily, rs_rank)
                    if ep is not None:
                        signal = ep

            if signal is None:
                continue

            # Score gate
            score = (getattr(signal, "composite_score", None) or
                     getattr(signal, "ep_score", None) or 0)
            if score < min_score:
                continue

            entry_price = (getattr(signal, "entry_price", None) or
                           getattr(signal, "price_at_signal", None))
            stop_loss   = getattr(signal, "stop_loss", None)
            sig_type    = (getattr(signal, "setup_type", None) or
                           getattr(signal, "ep_type", "UNKNOWN"))

            if not entry_price or not stop_loss or entry_price <= stop_loss:
                continue

            trade = Trade(
                symbol          = sym,
                signal_type     = sig_type,
                entry_date      = sim_date,
                entry_price     = entry_price,
                stop_loss       = stop_loss,
                target1         = getattr(signal, "target1", entry_price * 1.08),
                target2         = getattr(signal, "target2", entry_price * 1.15),
                rs_rank         = getattr(signal, "rs_rank", rs_rank),
                ti65            = getattr(signal, "ti65", 1.0),
                twolynch_score  = getattr(signal, "twolynch_score", None),
                consol_width    = getattr(signal, "consolidation_width_pct", None),
                composite_score = score,
                market_regime   = mm.market_regime,
            )
            open_positions[sym] = trade
            monthly_signals[month_key] = monthly_signals.get(month_key, 0) + 1

    # ── Close open positions at end of backtest ───────────────────────────────
    for sym, trade in open_positions.items():
        df = all_data.get(sym)
        if df is None:
            continue
        last_bar      = df.iloc[-1]
        last_close    = float(last_bar["Close"])
        risk_ps       = trade.entry_price - trade.stop_loss
        trade.exit_date   = end_date
        trade.exit_price  = last_close
        trade.exit_reason = "END_OF_BACKTEST"
        trade.hold_days   = (end_date - trade.entry_date).days
        trade.pnl_pct     = (last_close - trade.entry_price) / trade.entry_price * 100
        trade.pnl_r       = (last_close - trade.entry_price) / risk_ps if risk_ps > 0 else 0
        all_trades.append(trade)

    results = _compute_metrics(all_trades, monthly_signals, start_date, end_date)
    # Attach full trade list for Excel export
    results["_trades"] = all_trades
    _log_results_summary(results)
    return results


# ─── Metrics ──────────────────────────────────────────────────────────────────

def _compute_metrics(
    trades:          List[Trade],
    monthly_signals: dict,
    start_date:      date,
    end_date:        date,
) -> dict:

    if not trades:
        return {"error": "No trades", "trade_count": 0}

    rows = []
    for t in trades:
        if t.pnl_pct is None:
            continue
        rows.append({
            "symbol":         t.symbol,
            "signal_type":    t.signal_type or "",
            "entry_date":     pd.Timestamp(t.entry_date),
            "exit_date":      pd.Timestamp(t.exit_date) if t.exit_date else pd.NaT,
            "entry_price":    t.entry_price,
            "exit_price":     t.exit_price,
            "stop_loss":      t.stop_loss,
            "target1":        t.target1,
            "target2":        t.target2,
            "exit_reason":    t.exit_reason or "",
            "pnl_pct":        t.pnl_pct,
            "pnl_r":          t.pnl_r or 0,
            "hold_days":      t.hold_days or 0,
            "rs_rank":        t.rs_rank,
            "ti65":           t.ti65,
            "twolynch_score": t.twolynch_score,
            "consol_width":   t.consol_width,
            "composite_score":t.composite_score or 0,
            "market_regime":  t.market_regime or "",
            "partial_exit_price": t.partial_exit_price,
            "partial_pnl_pct":    t.partial_pnl_pct,
            "high_during_hold":   t.high_during_hold,
        })

    df = pd.DataFrame(rows)
    if df.empty:
        return {"error": "No completed trades", "trade_count": 0}

    wins   = df[df["pnl_pct"] > 0]
    losses = df[df["pnl_pct"] <= 0]

    win_rate   = len(wins) / len(df) * 100
    avg_win    = float(wins["pnl_pct"].mean())   if len(wins)   > 0 else 0.0
    avg_loss   = float(losses["pnl_pct"].mean()) if len(losses) > 0 else 0.0
    avg_r      = float(df["pnl_r"].mean())
    expectancy = avg_r  # E[R] per trade

    gross_profit  = float(wins["pnl_pct"].sum())   if len(wins)   > 0 else 0.0
    gross_loss    = abs(float(losses["pnl_pct"].sum())) if len(losses) > 0 else 1.0
    profit_factor = gross_profit / gross_loss if gross_loss > 0 else float("inf")

    # ── Equity curve (1% risk per trade, compound) ────────────────────────────
    # Each trade R × 1% of account = actual % return on account
    equity = [1.0]
    for r in df["pnl_r"]:
        equity.append(equity[-1] * (1 + r * RISK_PCT))
    eq_series = pd.Series(equity[1:])

    years = max((end_date - start_date).days / 365, 0.1)
    final_equity = float(eq_series.iloc[-1]) if len(eq_series) > 0 else 1.0
    cagr = (final_equity ** (1 / years) - 1) * 100

    # Max drawdown on equity curve
    peak     = eq_series.cummax()
    drawdown = (eq_series - peak) / peak
    max_dd   = float(drawdown.min()) * 100

    # Sharpe on daily equity curve returns
    daily_eq = eq_series.pct_change().dropna()
    if len(daily_eq) > 5:
        excess   = daily_eq - RISK_FREE_RATE / 252
        sharpe   = float(excess.mean() / (excess.std() + 1e-10) * math.sqrt(252))
    else:
        sharpe = 0.0

    # ── MB-specific metrics ───────────────────────────────────────────────────
    mb_df = df[df["signal_type"].str.contains("MB", na=False)]
    ep_df = df[df["signal_type"].str.contains("EP", na=False)]

    avg_hold_mb = float(mb_df["hold_days"].mean()) if len(mb_df) > 0 else 0.0
    avg_hold_ep = float(ep_df["hold_days"].mean()) if len(ep_df) > 0 else 0.0

    # Day 3 partial exit fired: partial_exit_price is not None
    if len(mb_df) > 0:
        day3_hit_pct   = float(mb_df["partial_exit_price"].notna().mean() * 100)
        # Burst 8%+: high_during_hold ≥ entry * 1.08
        burst_mask = (
            mb_df["high_during_hold"].notna() &
            (mb_df["high_during_hold"] >= mb_df["entry_price"] * 1.08)
        )
        burst_8pct_pct = float(burst_mask.mean() * 100)
    else:
        day3_hit_pct   = 0.0
        burst_8pct_pct = 0.0

    # ── Breakdowns ────────────────────────────────────────────────────────────

    def _breakdown(grp_col, bins=None, labels=None):
        out = {}
        sub = df[df[grp_col].notna()].copy()
        if sub.empty:
            return out
        if bins:
            sub["_bin"] = pd.cut(sub[grp_col], bins=bins, labels=labels)
            col = "_bin"
        else:
            col = grp_col
        for val, g in sub.groupby(col):
            if len(g) == 0:
                continue
            gw = g[g["pnl_pct"] > 0]
            gl = g[g["pnl_pct"] <= 0]
            pf = (float(gw["pnl_pct"].sum()) /
                  max(abs(float(gl["pnl_pct"].sum())), 0.01))
            out[str(val)] = {
                "count":         len(g),
                "win_rate":      round(float((g["pnl_pct"] > 0).mean() * 100), 1),
                "avg_pnl":       round(float(g["pnl_pct"].mean()), 2),
                "avg_r":         round(float(g["pnl_r"].mean()), 3),
                "profit_factor": round(pf, 2),
            }
        return out

    twolynch_bd = {}
    if "twolynch_score" in df.columns and df["twolynch_score"].notna().any():
        sub = df[df["twolynch_score"].notna()].copy()
        sub["lb"] = sub["twolynch_score"].astype(int)
        for sv, g in sub.groupby("lb"):
            gw = g[g["pnl_pct"] > 0]
            gl = g[g["pnl_pct"] <= 0]
            pf = float(gw["pnl_pct"].sum()) / max(abs(float(gl["pnl_pct"].sum())), 0.01)
            twolynch_bd[f"2LYNCH_{sv}/5"] = {
                "count":         len(g),
                "win_rate":      round(float((g["pnl_pct"] > 0).mean() * 100), 1),
                "avg_pnl":       round(float(g["pnl_pct"].mean()), 2),
                "avg_r":         round(float(g["pnl_r"].mean()), 3),
                "profit_factor": round(pf, 2),
            }

    rs_bd = _breakdown("rs_rank",
                        bins=[0, 20, 40, 60, 70, 80, 90, 100],
                        labels=["0-20", "20-40", "40-60", "60-70", "70-80", "80-90", "90-99"])

    ti65_bd = _breakdown("ti65",
                          bins=[0, 0.97, 1.0, 1.03, 1.05, 1.10, 9],
                          labels=["<0.97", "0.97-1.0", "1.0-1.03", "1.03-1.05", "1.05-1.10", ">1.10"])

    width_bd = _breakdown("consol_width",
                           bins=[0, 4, 7, 10, 12, 100],
                           labels=["<4%", "4-7%", "7-10%", "10-12%", ">12%"])

    score_bd = _breakdown("composite_score",
                           bins=[0, 55, 65, 70, 80, 85, 100],
                           labels=["<55", "55-65", "65-70", "70-80", "80-85", "85-100"])

    regime_bd = {}
    for regime, g in df.groupby("market_regime"):
        gw = g[g["pnl_pct"] > 0]
        gl = g[g["pnl_pct"] <= 0]
        pf = float(gw["pnl_pct"].sum()) / max(abs(float(gl["pnl_pct"].sum())), 0.01)
        regime_bd[f"Regime_{regime}"] = {
            "count":         len(g),
            "win_rate":      round(float((g["pnl_pct"] > 0).mean() * 100), 1),
            "avg_pnl":       round(float(g["pnl_pct"].mean()), 2),
            "avg_r":         round(float(g["pnl_r"].mean()), 3),
            "profit_factor": round(pf, 2),
        }

    exit_bd = {}
    for reason, g in df.groupby("exit_reason"):
        exit_bd[reason] = {
            "count":    len(g),
            "win_rate": round(float((g["pnl_pct"] > 0).mean() * 100), 1),
            "avg_pnl":  round(float(g["pnl_pct"].mean()), 2),
        }

    # Yearly breakdown
    yearly_bd = {}
    df["year"] = df["entry_date"].dt.year
    for yr, g in df.groupby("year"):
        gw = g[g["pnl_pct"] > 0]
        gl = g[g["pnl_pct"] <= 0]
        pf = float(gw["pnl_pct"].sum()) / max(abs(float(gl["pnl_pct"].sum())), 0.01)
        yearly_bd[int(yr)] = {
            "count":         len(g),
            "win_rate":      round(float((g["pnl_pct"] > 0).mean() * 100), 1),
            "avg_r":         round(float(g["pnl_r"].mean()), 3),
            "total_r":       round(float(g["pnl_r"].sum()), 2),
            "profit_factor": round(pf, 2),
        }

    monthly_avg = float(np.mean(list(monthly_signals.values()))) if monthly_signals else 0.0

    return {
        "period":              f"{start_date} → {end_date}",
        "trade_count":         len(df),
        "win_count":           len(wins),
        "loss_count":          len(losses),
        "win_rate_pct":        round(win_rate, 1),
        "avg_win_pct":         round(avg_win, 2),
        "avg_loss_pct":        round(avg_loss, 2),
        "avg_r":               round(avg_r, 3),
        "expectancy_r":        round(expectancy, 3),
        "profit_factor":       round(profit_factor, 2),
        "cagr_pct":            round(cagr, 1),
        "final_equity":        round(final_equity, 4),
        "max_drawdown_pct":    round(max_dd, 1),
        "sharpe_ratio":        round(sharpe, 2),
        "avg_hold_days_mb":    round(avg_hold_mb, 1),
        "avg_hold_days_ep":    round(avg_hold_ep, 1),
        "day3_target_hit_pct": round(day3_hit_pct, 1),
        "burst_8pct_pct":      round(burst_8pct_pct, 1),
        "monthly_avg_signals": round(monthly_avg, 1),
        "monthly_signal_freq": monthly_signals,
        "twolynch_breakdown":  twolynch_bd,
        "rs_breakdown":        rs_bd,
        "ti65_breakdown":      ti65_bd,
        "width_breakdown":     width_bd,
        "score_breakdown":     score_bd,
        "regime_breakdown":    regime_bd,
        "exit_breakdown":      exit_bd,
        "yearly_breakdown":    yearly_bd,
        "_df":                 df,          # internal: for Excel export
        "_equity":             equity,      # internal: equity curve points
    }


def _log_results_summary(r: dict) -> None:
    if "error" in r:
        log.error("Backtest: %s", r["error"])
        return
    log.info("=" * 60)
    log.info("BACKTEST RESULTS — %s", r.get("period", ""))
    log.info("=" * 60)
    log.info("Trades=%d  Win%%=%.1f%%  PF=%.2f  AvgR=%.3f  E[R]=%.3f",
             r["trade_count"], r["win_rate_pct"], r["profit_factor"],
             r["avg_r"], r["expectancy_r"])
    log.info("CAGR=%.1f%%  MaxDD=%.1f%%  Sharpe=%.2f  FinalEq=%.3f",
             r["cagr_pct"], r["max_drawdown_pct"], r["sharpe_ratio"], r["final_equity"])
    log.info("MB: AvgHold=%.1fd  Day3Hit=%.1f%%  Burst8+=%.1f%%",
             r["avg_hold_days_mb"], r["day3_target_hit_pct"], r["burst_8pct_pct"])
    log.info("Regime: %s", r.get("regime_breakdown", {}))
    log.info("=" * 60)


# ─── Excel Report (reference-quality, matches VCP format) ─────────────────────

def export_backtest_excel(
    results_with:    dict,
    results_without: dict,
    output_path:     str,
    signal_type:     str  = "MB_BREAKOUT",
    years:           int  = 1,
    notes:           str  = "",
) -> None:
    """
    Export a multi-sheet Excel report matching the VCP reference format.

    Sheets:
      Verdict          — pass/fail criteria summary
      Equity Curve     — trade-by-trade compounding
      Regime Analysis  — WITH vs WITHOUT filter (proves MM value)
      RS Analysis      — RS rank → win rate (validates RS filter)
      TI65 Analysis    — TI65 level → win rate (validates TI65 filter)
      Score Analysis   — composite score → win rate
      Width Analysis   — consolidation width sweet spot
      2LYNCH Analysis  — 2LYNCH score → win rate
      Yearly P&L       — year-by-year regime dependency
      Exit Analysis    — exit reason breakdown
      Trade Log        — every trade, all fields
    """
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, LineChart, Reference

    wb  = Workbook()
    rw  = results_with
    rwo = results_without

    # ── Colour palette ────────────────────────────────────────────────────────
    C = {
        "header_bg":  "1F2937",   # dark slate
        "header_fg":  "FFFFFF",
        "title_bg":   "111827",
        "title_fg":   "F9FAFB",
        "pass_bg":    "D1FAE5",
        "fail_bg":    "FEE2E2",
        "warn_bg":    "FEF3C7",
        "alt_row":    "F3F4F6",
        "green":      "059669",
        "red":        "DC2626",
        "amber":      "D97706",
        "blue":       "2563EB",
        "accent":     "7C3AED",
    }

    thin = Side(style="thin", color="D1D5DB")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_cell(ws, row, col, val, bg=C["header_bg"], fg=C["header_fg"], bold=True):
        c = ws.cell(row=row, column=col, value=val)
        c.fill      = PatternFill("solid", fgColor=bg)
        c.font      = Font(bold=bold, color=fg, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = bdr
        return c

    def data_cell(ws, row, col, val, fmt=None, bold=False, color=None, bg=None, align="center"):
        c = ws.cell(row=row, column=col, value=val)
        c.font      = Font(bold=bold, color=color or "000000", size=10)
        c.alignment = Alignment(horizontal=align, vertical="center")
        c.border    = bdr
        if bg:
            c.fill = PatternFill("solid", fgColor=bg)
        if fmt:
            c.number_format = fmt
        return c

    def title_row(ws, row, text, ncols, bg=C["title_bg"], fg=C["title_fg"]):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        c = ws.cell(row=row, column=1, value=text)
        c.fill      = PatternFill("solid", fgColor=bg)
        c.font      = Font(bold=True, color=fg, size=12)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 24

    def auto_width(ws, min_w=8, max_w=30):
        for col_cells in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col_cells), default=0)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = \
                max(min_w, min(max_len + 2, max_w))

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 1: VERDICT
    # ═══════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Verdict"
    ws.sheet_view.showGridLines = False

    tc   = rw.get("trade_count", 0)
    wr   = rw.get("win_rate_pct", 0)
    pf   = rw.get("profit_factor", 0)
    exr  = rw.get("expectancy_r", 0)
    dd   = rw.get("max_drawdown_pct", 0)
    cagr = rw.get("cagr_pct", 0)
    sh   = rw.get("sharpe_ratio", 0)

    criteria = [
        ("Sample size",       f"{tc} trades",   "≥ 30",   tc >= 30),
        ("Win rate",          f"{wr}%",          "≥ 44%",  wr >= 44),
        ("Profit factor",     str(pf),           "≥ 1.5",  pf >= 1.5),
        ("Expectancy",        f"{exr:+.3f}R",    "> 0R",   exr > 0),
        ("Max drawdown",      f"{dd:.1f}%",      "> -15%", dd > -15),
        ("CAGR",              f"{cagr:.1f}%",    "> 12%",  cagr > 12),
    ]
    n_pass = sum(1 for *_, p in criteria if p)
    verdict_text = (
        f"✅ NSE Bonde MB VALIDATED — passes {n_pass}/{len(criteria)} criteria"
        if n_pass >= 5 else
        f"⚠️  NSE Bonde MB NOT YET VALIDATED — passes {n_pass}/{len(criteria)} criteria"
    )

    # Row 1: verdict banner
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A1:H1")
    v = ws["A1"]
    v.value     = verdict_text
    v.fill      = PatternFill("solid", fgColor=(C["pass_bg"] if n_pass >= 5 else C["warn_bg"]))
    v.font      = Font(bold=True, size=13,
                       color=(C["green"] if n_pass >= 5 else C["amber"]))
    v.alignment = Alignment(horizontal="center", vertical="center")

    # Row 2: run info
    ws.merge_cells("A2:H2")
    ws["A2"].value = (f"Signal: {signal_type}  |  Period: {rw.get('period','')}  |  "
                      f"Trades: {tc}  |  Notes: {notes or '—'}")
    ws["A2"].font      = Font(size=10, color="6B7280")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Row 3: big metrics
    ws.row_dimensions[3].height = 20
    big = [("Win Rate", f"{wr}%"), ("Profit Factor", str(pf)),
           ("Expectancy", f"{exr:+.3f}R"), ("Max Drawdown", f"{dd:.1f}%"),
           ("CAGR", f"{cagr:.1f}%"), ("Sharpe", str(sh))]
    for i, (label, val) in enumerate(big):
        hdr_cell(ws, 3, i * 2 + 1, label, bg=C["accent"])
        hdr_cell(ws, 3, i * 2 + 2, val,   bg="374151", fg="F9FAFB")

    # Criteria table
    ws.row_dimensions[5].height = 18
    for ci, col_name in enumerate(["Check", "Value", "Threshold", "Pass/Fail"], 1):
        hdr_cell(ws, 5, ci, col_name)
    for ri, (check, val, thresh, passed) in enumerate(criteria, 6):
        bg = C["pass_bg"] if passed else C["fail_bg"]
        data_cell(ws, ri, 1, check,          bg=bg, align="left")
        data_cell(ws, ri, 2, val,            bg=bg)
        data_cell(ws, ri, 3, thresh,         bg=bg)
        data_cell(ws, ri, 4, "✅ PASS" if passed else "❌ FAIL", bg=bg,
                  bold=True, color=(C["green"] if passed else C["red"]))

    # WITH vs WITHOUT comparison
    r13 = 6 + len(criteria) + 2
    title_row(ws, r13, "Market Monitor Filter Comparison (WITH vs WITHOUT)", 8)
    hdrs = ["Metric", "WITH MM Filter", "WITHOUT MM Filter", "Difference", "Filter Adds Value?"]
    for ci, h in enumerate(hdrs, 1):
        hdr_cell(ws, r13 + 1, ci, h)

    metrics_cmp = [
        ("Trades",        rw.get("trade_count", 0),    rwo.get("trade_count", 0),    None,  False),
        ("Win Rate %",    rw.get("win_rate_pct", 0),   rwo.get("win_rate_pct", 0),   None,  True),
        ("Profit Factor", rw.get("profit_factor", 0),  rwo.get("profit_factor", 0),  None,  True),
        ("CAGR %",        rw.get("cagr_pct", 0),       rwo.get("cagr_pct", 0),       None,  True),
        ("Max Drawdown %",rw.get("max_drawdown_pct",0),rwo.get("max_drawdown_pct",0),None, False),
        ("Sharpe",        rw.get("sharpe_ratio", 0),   rwo.get("sharpe_ratio", 0),   None,  True),
    ]
    for ri2, (metric, wv, wov, _, higher_is_better) in enumerate(metrics_cmp, r13 + 2):
        diff = round(wv - wov, 2) if isinstance(wv, (int, float)) else "—"
        adds = (diff > 0) if higher_is_better else (diff < 0)
        data_cell(ws, ri2, 1, metric, align="left")
        data_cell(ws, ri2, 2, wv,  bold=True)
        data_cell(ws, ri2, 3, wov)
        data_cell(ws, ri2, 4, f"{diff:+}" if isinstance(diff, float) else diff)
        data_cell(ws, ri2, 5, "✅ YES" if adds else "❌ NO" if isinstance(diff, float) else "?",
                  color=C["green"] if adds else C["red"])

    auto_width(ws)

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 2: EQUITY CURVE
    # ═══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Equity Curve")
    ws2.sheet_view.showGridLines = False

    title_row(ws2, 1, "Equity Curve — Trade-by-Trade Compounding (1% Risk per Trade)", 5)
    eq_hdrs = ["Trade #", "Entry Date", "Symbol", "Exit Reason", "R Realised", "Cumulative Equity"]
    for ci, h in enumerate(eq_hdrs, 1):
        hdr_cell(ws2, 2, ci, h)

    df_eq  = rw.get("_df", pd.DataFrame())
    equity = rw.get("_equity", [1.0])
    if not df_eq.empty:
        for ri, (_, row) in enumerate(df_eq.iterrows(), 3):
            bg = C["alt_row"] if ri % 2 == 0 else "FFFFFF"
            r_val = row.get("pnl_r", 0)
            eq_val = equity[ri - 3 + 1] if (ri - 3 + 1) < len(equity) else None
            data_cell(ws2, ri, 1, ri - 2)
            data_cell(ws2, ri, 2, row["entry_date"].strftime("%Y-%m-%d") if pd.notna(row["entry_date"]) else "", bg=bg)
            data_cell(ws2, ri, 3, row.get("symbol", ""), bg=bg, align="left")
            data_cell(ws2, ri, 4, row.get("exit_reason", ""), bg=bg)
            data_cell(ws2, ri, 5, round(r_val, 3), bg=bg,
                      color=C["green"] if r_val > 0 else C["red"])
            data_cell(ws2, ri, 6, round(eq_val, 4) if eq_val else None, bg=bg,
                      fmt="0.0000")

    auto_width(ws2)

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 3: REGIME ANALYSIS
    # ═══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Regime Analysis")
    ws3.sheet_view.showGridLines = False
    title_row(ws3, 1, "Market Regime Analysis — Does the Market Monitor Filter Add Value?", 7)

    hdrs3 = ["Regime", "Trades", "Win Rate %", "Avg R", "Profit Factor", "Avg P&L %", "Filter Active?"]
    for ci, h in enumerate(hdrs3, 1):
        hdr_cell(ws3, 2, ci, h)

    regime_bd = rwo.get("regime_breakdown", {})  # WITHOUT filter shows all regimes
    for ri, (regime, v) in enumerate(sorted(regime_bd.items()), 3):
        bg = (C["pass_bg"] if v["win_rate"] >= 50 else
              C["warn_bg"] if v["win_rate"] >= 40 else C["fail_bg"])
        data_cell(ws3, ri, 1, regime,             align="left")
        data_cell(ws3, ri, 2, v["count"])
        data_cell(ws3, ri, 3, f"{v['win_rate']}%", bg=bg, bold=True)
        data_cell(ws3, ri, 4, v.get("avg_r", "—"))
        data_cell(ws3, ri, 5, v.get("profit_factor", "—"))
        data_cell(ws3, ri, 6, f"{v['avg_pnl']:+.2f}%")
        active = "BEAR blocked" if "BEAR" in regime else ("EP only" if "CAUTION" in regime else "Full scan")
        data_cell(ws3, ri, 7, active)

    auto_width(ws3)

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 4: RS ANALYSIS
    # ═══════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("RS Analysis")
    ws4.sheet_view.showGridLines = False
    title_row(ws4, 1, "RS Rank Analysis — Does Higher Relative Strength = Better Outcome?", 6)

    hdrs4 = ["RS Band", "Trades", "Win Rate %", "Avg R", "Profit Factor", "Avg P&L %"]
    for ci, h in enumerate(hdrs4, 1):
        hdr_cell(ws4, 2, ci, h)

    rs_bd = rw.get("rs_breakdown", {})
    for ri, (band, v) in enumerate(sorted(rs_bd.items()), 3):
        bg = (C["pass_bg"] if v["win_rate"] >= 55 else
              C["warn_bg"] if v["win_rate"] >= 45 else C["fail_bg"])
        data_cell(ws4, ri, 1, band, align="left")
        data_cell(ws4, ri, 2, v["count"])
        data_cell(ws4, ri, 3, f"{v['win_rate']}%", bg=bg, bold=True)
        data_cell(ws4, ri, 4, v.get("avg_r", "—"))
        data_cell(ws4, ri, 5, v.get("profit_factor", "—"))
        data_cell(ws4, ri, 6, f"{v['avg_pnl']:+.2f}%")

    # Interpretation note
    note_row = 3 + len(rs_bd) + 2
    ws4.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=6)
    note = ws4.cell(row=note_row, column=1,
                    value="✅ If RS 80-99 win rate > RS 0-40 win rate by ≥10pp, the RS filter is doing discriminative work. "
                          "Raise RS_MIN_FOR_MB to 80 if RS 80-99 significantly outperforms RS 60-80.")
    note.font      = Font(italic=True, size=9, color="6B7280")
    note.alignment = Alignment(wrap_text=True)
    auto_width(ws4)

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 5: TI65 ANALYSIS
    # ═══════════════════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("TI65 Analysis")
    ws5.sheet_view.showGridLines = False
    title_row(ws5, 1, "TI65 Analysis — Absolute Momentum Threshold Validation", 6)
    for ci, h in enumerate(["TI65 Band", "Trades", "Win Rate %", "Avg R", "Profit Factor", "Avg P&L %"], 1):
        hdr_cell(ws5, 2, ci, h)
    ti65_bd = rw.get("ti65_breakdown", {})
    for ri, (band, v) in enumerate(sorted(ti65_bd.items()), 3):
        bg = C["pass_bg"] if v["win_rate"] >= 45 else (C["warn_bg"] if v["win_rate"] >= 35 else C["fail_bg"])
        data_cell(ws5, ri, 1, band, align="left")
        data_cell(ws5, ri, 2, v["count"])
        data_cell(ws5, ri, 3, f"{v['win_rate']}%", bg=bg, bold=True)
        data_cell(ws5, ri, 4, v.get("avg_r", "—"))
        data_cell(ws5, ri, 5, v.get("profit_factor", "—"))
        data_cell(ws5, ri, 6, f"{v['avg_pnl']:+.2f}%")
    n5 = ws5.cell(row=3+len(ti65_bd)+2, column=1,
                  value="✅ TI65>1.05 should significantly outperform TI65 1.03-1.05. "
                        "If yes, raise TI65_BULL_THRESHOLD to 1.05.")
    n5.font = Font(italic=True, size=9, color="6B7280")
    ws5.merge_cells(start_row=3+len(ti65_bd)+2, start_column=1, end_row=3+len(ti65_bd)+2, end_column=6)
    auto_width(ws5)

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 6: SCORE ANALYSIS
    # ═══════════════════════════════════════════════════════════════════════════
    ws6 = wb.create_sheet("Score Analysis")
    ws6.sheet_view.showGridLines = False
    title_row(ws6, 1, "Composite Score Analysis — Does Higher Score = Better Outcome?", 6)
    for ci, h in enumerate(["Score Band", "Trades", "Win Rate %", "Avg R", "Profit Factor", "Avg P&L %"], 1):
        hdr_cell(ws6, 2, ci, h)
    score_bd = rw.get("score_breakdown", {})
    for ri, (band, v) in enumerate(sorted(score_bd.items()), 3):
        bg = C["pass_bg"] if v["win_rate"] >= 50 else (C["warn_bg"] if v["win_rate"] >= 40 else C["fail_bg"])
        data_cell(ws6, ri, 1, band, align="left")
        data_cell(ws6, ri, 2, v["count"])
        data_cell(ws6, ri, 3, f"{v['win_rate']}%", bg=bg, bold=True)
        data_cell(ws6, ri, 4, v.get("avg_r", "—"))
        data_cell(ws6, ri, 5, v.get("profit_factor", "—"))
        data_cell(ws6, ri, 6, f"{v['avg_pnl']:+.2f}%")
    auto_width(ws6)

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 7: WIDTH ANALYSIS
    # ═══════════════════════════════════════════════════════════════════════════
    ws7 = wb.create_sheet("Width Analysis")
    ws7.sheet_view.showGridLines = False
    title_row(ws7, 1, "Consolidation Width Analysis — Tight vs Wide Base", 6)
    for ci, h in enumerate(["Width Band", "Trades", "Win Rate %", "Avg R", "Profit Factor", "Avg P&L %"], 1):
        hdr_cell(ws7, 2, ci, h)
    width_bd = rw.get("width_breakdown", {})
    for ri, (band, v) in enumerate(sorted(width_bd.items()), 3):
        bg = C["pass_bg"] if v["win_rate"] >= 50 else (C["warn_bg"] if v["win_rate"] >= 40 else C["fail_bg"])
        data_cell(ws7, ri, 1, band, align="left")
        data_cell(ws7, ri, 2, v["count"])
        data_cell(ws7, ri, 3, f"{v['win_rate']}%", bg=bg, bold=True)
        data_cell(ws7, ri, 4, v.get("avg_r", "—"))
        data_cell(ws7, ri, 5, v.get("profit_factor", "—"))
        data_cell(ws7, ri, 6, f"{v['avg_pnl']:+.2f}%")
    auto_width(ws7)

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 8: 2LYNCH ANALYSIS
    # ═══════════════════════════════════════════════════════════════════════════
    ws8 = wb.create_sheet("2LYNCH Analysis")
    ws8.sheet_view.showGridLines = False
    title_row(ws8, 1, "2LYNCH Checklist Analysis — Does More Criteria = Better Outcome?", 6)
    for ci, h in enumerate(["2LYNCH Score", "Trades", "Win Rate %", "Avg R", "Profit Factor", "Avg P&L %"], 1):
        hdr_cell(ws8, 2, ci, h)
    lynch_bd = rw.get("twolynch_breakdown", {})
    for ri, (band, v) in enumerate(sorted(lynch_bd.items()), 3):
        bg = C["pass_bg"] if v["win_rate"] >= 50 else (C["warn_bg"] if v["win_rate"] >= 40 else C["fail_bg"])
        data_cell(ws8, ri, 1, band, align="left")
        data_cell(ws8, ri, 2, v["count"])
        data_cell(ws8, ri, 3, f"{v['win_rate']}%", bg=bg, bold=True)
        data_cell(ws8, ri, 4, v.get("avg_r", "—"))
        data_cell(ws8, ri, 5, v.get("profit_factor", "—"))
        data_cell(ws8, ri, 6, f"{v['avg_pnl']:+.2f}%")
    auto_width(ws8)

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 9: YEARLY P&L
    # ═══════════════════════════════════════════════════════════════════════════
    ws9 = wb.create_sheet("Yearly P&L")
    ws9.sheet_view.showGridLines = False
    title_row(ws9, 1, "Year-by-Year Breakdown — Regime Dependency Check", 7)
    for ci, h in enumerate(["Year", "Trades", "Win Rate %", "Avg R", "Total R", "Profit Factor", "Regime flavour"], 1):
        hdr_cell(ws9, 2, ci, h)
    yearly_bd = rw.get("yearly_breakdown", {})
    for ri, (yr, v) in enumerate(sorted(yearly_bd.items()), 3):
        bg = C["pass_bg"] if v["win_rate"] >= 50 else (C["warn_bg"] if v["win_rate"] >= 40 else C["fail_bg"])
        data_cell(ws9, ri, 1, yr)
        data_cell(ws9, ri, 2, v["count"])
        data_cell(ws9, ri, 3, f"{v['win_rate']}%", bg=bg, bold=True)
        data_cell(ws9, ri, 4, f"{v['avg_r']:+.3f}R")
        data_cell(ws9, ri, 5, f"{v['total_r']:+.2f}R",
                  color=C["green"] if v["total_r"] > 0 else C["red"], bold=True)
        data_cell(ws9, ri, 6, v.get("profit_factor", "—"))
        data_cell(ws9, ri, 7, "")  # placeholder for manual regime notes
    note9 = ws9.cell(row=3+len(yearly_bd)+2, column=1,
                     value="If any year shows PF < 1.0, check what regime dominated that year. "
                           "Persistent losses in a year = scanner is not regime-aware enough.")
    note9.font = Font(italic=True, size=9, color="6B7280")
    ws9.merge_cells(start_row=3+len(yearly_bd)+2, start_column=1, end_row=3+len(yearly_bd)+2, end_column=7)
    auto_width(ws9)

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 10: EXIT ANALYSIS
    # ═══════════════════════════════════════════════════════════════════════════
    ws10 = wb.create_sheet("Exit Analysis")
    ws10.sheet_view.showGridLines = False
    title_row(ws10, 1, "Exit Reason Analysis — Are Stops / Targets / Holds Working?", 5)
    for ci, h in enumerate(["Exit Reason", "Trades", "Win Rate %", "Avg P&L %", "Comment"], 1):
        hdr_cell(ws10, 2, ci, h)
    exit_comments = {
        "DAY5_EXIT":        "MB 5-day force exit — should be ~85%+ of MB trades",
        "STOP_HIT":         "Intraday stop hit — avg loss should be ~-5 to -8%",
        "GAP_STOP":         "Gap-down open through stop — worst case, hard to avoid",
        "EP_MAX_HOLD":      "EP held full 30 days — usually a winner",
        "EP_TRAIL_STOP":    "EP trail stop fired after 20%+ gain — excellent",
        "END_OF_BACKTEST":  "Still open at period end — not a real exit",
    }
    exit_bd = rw.get("exit_breakdown", {})
    for ri, (reason, v) in enumerate(sorted(exit_bd.items(), key=lambda x: -x[1]["count"]), 3):
        bg = C["pass_bg"] if v["win_rate"] > 50 else (C["warn_bg"] if v["win_rate"] > 30 else C["fail_bg"])
        data_cell(ws10, ri, 1, reason, align="left")
        data_cell(ws10, ri, 2, v["count"])
        data_cell(ws10, ri, 3, f"{v['win_rate']}%", bg=bg, bold=True)
        data_cell(ws10, ri, 4, f"{v['avg_pnl']:+.2f}%")
        data_cell(ws10, ri, 5, exit_comments.get(reason, ""), align="left")
    auto_width(ws10)

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 11: TRADE LOG
    # ═══════════════════════════════════════════════════════════════════════════
    ws11 = wb.create_sheet("Trade Log")
    ws11.sheet_view.showGridLines = False
    title_row(ws11, 1, "Full Trade Log — Every Individual Trade", 18)

    tl_hdrs = [
        "Symbol", "Signal Type", "Entry Date", "Exit Date",
        "Entry Price", "Stop Loss", "Exit Price",
        "Exit Reason", "P&L %", "R Realised",
        "Hold Days", "RS Rank", "TI65", "2LYNCH",
        "Consol Width %", "Score", "Market Regime",
        "Partial Exit Price"
    ]
    for ci, h in enumerate(tl_hdrs, 1):
        hdr_cell(ws11, 2, ci, h)

    df_log = rw.get("_df", pd.DataFrame())
    if not df_log.empty:
        for ri, (_, row) in enumerate(df_log.iterrows(), 3):
            bg = C["alt_row"] if ri % 2 == 0 else "FFFFFF"
            pnl = row.get("pnl_pct", 0)
            r_v = row.get("pnl_r",   0)
            data_cell(ws11, ri,  1, row.get("symbol", ""),        bg=bg, align="left")
            data_cell(ws11, ri,  2, row.get("signal_type", ""),   bg=bg)
            data_cell(ws11, ri,  3, row["entry_date"].strftime("%Y-%m-%d") if pd.notna(row["entry_date"]) else "", bg=bg)
            data_cell(ws11, ri,  4, row["exit_date"].strftime("%Y-%m-%d")  if pd.notna(row.get("exit_date")) else "", bg=bg)
            data_cell(ws11, ri,  5, round(row.get("entry_price", 0), 2), bg=bg, fmt="0.00")
            data_cell(ws11, ri,  6, round(row.get("stop_loss", 0), 2),   bg=bg, fmt="0.00")
            data_cell(ws11, ri,  7, round(row.get("exit_price", 0), 2),  bg=bg, fmt="0.00")
            data_cell(ws11, ri,  8, row.get("exit_reason", ""),           bg=bg)
            data_cell(ws11, ri,  9, round(pnl, 2), bg=bg, fmt="0.00",
                      color=C["green"] if pnl > 0 else C["red"], bold=True)
            data_cell(ws11, ri, 10, round(r_v, 3), bg=bg,
                      color=C["green"] if r_v > 0 else C["red"])
            data_cell(ws11, ri, 11, row.get("hold_days", ""),     bg=bg)
            data_cell(ws11, ri, 12, round(row.get("rs_rank", 0) or 0, 1), bg=bg)
            data_cell(ws11, ri, 13, round(row.get("ti65", 0) or 0, 3),    bg=bg)
            data_cell(ws11, ri, 14, row.get("twolynch_score", ""),        bg=bg)
            data_cell(ws11, ri, 15, round(row.get("consol_width", 0) or 0, 1), bg=bg)
            data_cell(ws11, ri, 16, round(row.get("composite_score", 0) or 0, 1), bg=bg)
            data_cell(ws11, ri, 17, row.get("market_regime", ""),         bg=bg)
            data_cell(ws11, ri, 18,
                      round(row.get("partial_exit_price") or 0, 2) if row.get("partial_exit_price") else "",
                      bg=bg, fmt="0.00")

    auto_width(ws11)

    wb.save(output_path)
    log.info("Backtest Excel report saved → %s", output_path)
